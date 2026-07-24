"""
Excel report generator using openpyxl.

Produces a workbook with five sheets:
  1. Overview   – KPI metrics, quality gate, ratings
  2. Issues     – Full issue list with filters and conditional formatting
  3. Files      – Per-file aggregated statistics
  4. Rules      – Per-rule violation counts
  5. Metrics    – All raw measures from SonarQube
"""

from __future__ import annotations

import io
import logging
from typing import TYPE_CHECKING, Any

import openpyxl
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from models.issue import IssueType, Severity
from models.metrics import rating_color, rating_letter
from models.project import ReportData
from reports.base_report import BaseReport

if TYPE_CHECKING:
    pass

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colour constants (Excel ARGB hex)
# ---------------------------------------------------------------------------

C_PRIMARY    = "FF1A237E"
C_SECONDARY  = "FF283593"
C_ACCENT     = "FF42A5F5"
C_SUCCESS    = "FF4CAF50"
C_WARNING    = "FFFF9800"
C_DANGER     = "FFF44336"
C_ALT_ROW    = "FFF0F4FF"
C_WHITE      = "FFFFFFFF"
C_HEADER_FG  = "FFFFFFFF"
C_DARK_TEXT  = "FF212121"

SEV_COLORS = {
    "BLOCKER":  "FFD32F2F",
    "CRITICAL": "FFF44336",
    "MAJOR":    "FFFF9800",
    "MINOR":    "FFFFC107",
    "INFO":     "FF2196F3",
}

TYPE_COLORS = {
    "BUG":           "FFF44336",
    "VULNERABILITY": "FF9C27B0",
    "CODE_SMELL":    "FFFF9800",
}

RATING_FILLS = {
    "A": "FF4CAF50",
    "B": "FF8BC34A",
    "C": "FFFF9800",
    "D": "FFF44336",
    "E": "FFD32F2F",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _fill(hex_argb: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_argb)

def _font(bold: bool = False, color: str = C_DARK_TEXT, size: int = 10) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _border() -> Border:
    thin = Side(style="thin", color="FFD0D0D0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _header_row(ws, row: int, values: list[str], col_start: int = 1) -> None:
    """Write a styled header row."""
    for col, val in enumerate(values, start=col_start):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill   = _fill(C_PRIMARY)
        cell.font   = _font(bold=True, color=C_WHITE, size=10)
        cell.alignment = _center()
        cell.border = _border()


def _auto_width(ws, min_width: int = 8, max_width: int = 50) -> None:
    """Auto-fit column widths based on content."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 2, max_width))


# ---------------------------------------------------------------------------
# Excel Report
# ---------------------------------------------------------------------------

class ExcelReport(BaseReport):
    """Generates a multi-sheet Excel workbook."""

    def generate(self, data: ReportData) -> str:
        path = self.output_path("Report.xlsx")
        logger.info("Generating Excel report → %s", path)

        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        self._sheet_overview(wb, data)
        self._sheet_issues(wb, data)
        self._sheet_files(wb, data)
        self._sheet_rules(wb, data)
        self._sheet_metrics(wb, data)

        wb.save(path)
        logger.info("Excel report saved – %d sheets", len(wb.sheetnames))
        return path

    # ------------------------------------------------------------------
    # Sheet 1 – Overview
    # ------------------------------------------------------------------

    def _sheet_overview(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Overview")
        m  = data.metrics

        # Title
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value     = data.report_title
        title_cell.font      = _font(bold=True, color=C_WHITE, size=14)
        title_cell.fill      = _fill(C_PRIMARY)
        title_cell.alignment = _center()
        ws.row_dimensions[1].height = 30

        ws.merge_cells("A2:F2")
        sub = ws["A2"]
        sub.value     = (
            f"{data.company_name}  |  Generated: {self.format_datetime(data.generated_at)}"
            f"  |  Project: {data.project_key}"
        )
        sub.font      = _font(color=C_WHITE, size=9)
        sub.fill      = _fill(C_SECONDARY)
        sub.alignment = _center()

        # Section: Quality Gate
        self._write_section_title(ws, 4, "A", "Quality Gate")
        gate_val = "✅ PASSED" if m.quality_gate_status == "OK" else "❌ FAILED"
        gate_fill = C_SUCCESS if m.quality_gate_status == "OK" else C_DANGER
        ws["B5"].value = gate_val
        ws["B5"].font  = _font(bold=True, color=C_WHITE)
        ws["B5"].fill  = _fill(gate_fill)
        ws["B5"].alignment = _center()

        # Section: Key metrics
        self._write_section_title(ws, 7, "A", "Key Metrics")
        kpi_rows = [
            ("Lines of Code",          f"{m.ncloc:,}",              "–"),
            ("Files",                  str(m.files),                 "–"),
            ("Functions",              str(m.functions),             "–"),
            ("Bugs",                   str(m.bugs),                  m.reliability_rating_letter),
            ("Vulnerabilities",        str(m.vulnerabilities),       m.security_rating_letter),
            ("Code Smells",            str(m.code_smells),           m.sqale_rating_letter),
            ("Technical Debt",         m.technical_debt_display,     "–"),
            ("Coverage",               f"{m.coverage:.1f}%",         "–"),
            ("Duplicated Lines",       f"{m.duplicated_lines_density:.1f}%", "–"),
            ("Total Issues",           str(len(data.issues)),        "–"),
            ("Open Issues",            str(m.open_issues),           "–"),
            ("Accepted Issues",        str(m.accepted_issues),       "–"),
            ("Cyclomatic Complexity",  str(m.complexity),            "–"),
            ("Cognitive Complexity",   str(m.cognitive_complexity),  "–"),
        ]

        _header_row(ws, 8, ["Metric", "Value", "Rating"], col_start=1)
        for row_idx, (label, value, rating) in enumerate(kpi_rows, start=9):
            ws.cell(row=row_idx, column=1, value=label).font  = _font(bold=True)
            ws.cell(row=row_idx, column=2, value=value).alignment = _center()
            rc = ws.cell(row=row_idx, column=3, value=rating)
            rc.alignment = _center()
            if rating in RATING_FILLS:
                rc.fill = _fill(RATING_FILLS[rating])
                rc.font = _font(bold=True, color=C_WHITE)
            if row_idx % 2 == 0:
                ws.cell(row=row_idx, column=1).fill = _fill(C_ALT_ROW)
                ws.cell(row=row_idx, column=2).fill = _fill(C_ALT_ROW)

        for row in ws.iter_rows(min_row=8, max_row=8+len(kpi_rows)):
            for cell in row:
                cell.border = _border()

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        ws.freeze_panes = "A3"

    # ------------------------------------------------------------------
    # Sheet 2 – Issues
    # ------------------------------------------------------------------

    def _sheet_issues(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Issues")

        headers = ["#", "Rule", "Severity", "Type", "File", "Line",
                   "Status", "Effort", "Message", "Tags"]
        _header_row(ws, 1, headers)
        ws.freeze_panes = "A2"

        for idx, issue in enumerate(data.issues_sorted_by_severity, start=1):
            row = idx + 1
            values = [
                idx,
                issue.rule,
                issue.severity.value,
                issue.issue_type.label,
                issue.file_path,
                issue.display_line or "",
                issue.issue_status.value,
                issue.effort,
                issue.message,
                ", ".join(issue.tags),
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border    = _border()
                cell.alignment = _left()
                cell.font      = _font(size=9)
                if row % 2 == 0:
                    cell.fill = _fill(C_ALT_ROW)

            # Colour severity cell
            sev_cell = ws.cell(row=row, column=3)
            sev_fill = SEV_COLORS.get(issue.severity.value)
            if sev_fill:
                sev_cell.fill = _fill(sev_fill)
                sev_cell.font = _font(bold=True, color=C_WHITE, size=9)
            sev_cell.alignment = _center()

            # Colour type cell
            type_cell = ws.cell(row=row, column=4)
            type_fill = TYPE_COLORS.get(issue.issue_type.value)
            if type_fill:
                type_cell.fill = _fill(type_fill)
                type_cell.font = _font(bold=True, color=C_WHITE, size=9)

        # Auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        _auto_width(ws, max_width=60)
        ws.column_dimensions["E"].width = 50
        ws.column_dimensions["I"].width = 60

    # ------------------------------------------------------------------
    # Sheet 3 – Files
    # ------------------------------------------------------------------

    def _sheet_files(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Files")

        headers = ["Rank", "File Path", "Total Issues", "Bugs",
                   "Code Smells", "Vulnerabilities", "Critical",
                   "Major", "Effort (min)"]
        _header_row(ws, 1, headers)
        ws.freeze_panes = "A2"

        sorted_files = sorted(data.file_stats, key=lambda f: f.total_issues, reverse=True)
        for rank, fa in enumerate(sorted_files, start=1):
            row = rank + 1
            values = [
                rank, fa.path, fa.total_issues, fa.bugs,
                fa.code_smells, fa.vulnerabilities, fa.critical,
                fa.major, fa.effort_minutes,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border    = _border()
                cell.alignment = _left()
                cell.font      = _font(size=9)
                if row % 2 == 0:
                    cell.fill = _fill(C_ALT_ROW)

            # Heat-map on total issues column
            total_cell = ws.cell(row=row, column=3)
            total = fa.total_issues
            if total >= 20:
                total_cell.fill = _fill(C_DANGER)
                total_cell.font = _font(bold=True, color=C_WHITE, size=9)
            elif total >= 10:
                total_cell.fill = _fill(C_WARNING)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        _auto_width(ws, max_width=70)

    # ------------------------------------------------------------------
    # Sheet 4 – Rules
    # ------------------------------------------------------------------

    def _sheet_rules(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Rules")

        headers = ["Rank", "Rule Key", "Violations", "Severity",
                   "Type", "Effort (min)", "Sample Message"]
        _header_row(ws, 1, headers)
        ws.freeze_panes = "A2"

        sorted_rules = sorted(data.rule_stats, key=lambda r: r.count, reverse=True)
        for rank, ra in enumerate(sorted_rules, start=1):
            row = rank + 1
            values = [
                rank, ra.rule_key, ra.count, ra.severity,
                ra.issue_type.replace("_", " ").title(),
                ra.effort_minutes, ra.message_sample,
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border    = _border()
                cell.alignment = _left()
                cell.font      = _font(size=9)
                if row % 2 == 0:
                    cell.fill = _fill(C_ALT_ROW)

            # Colour severity cell
            sev_cell = ws.cell(row=row, column=4)
            sev_fill = SEV_COLORS.get(ra.severity)
            if sev_fill:
                sev_cell.fill = _fill(sev_fill)
                sev_cell.font = _font(bold=True, color=C_WHITE, size=9)
            sev_cell.alignment = _center()

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        _auto_width(ws, max_width=70)

    # ------------------------------------------------------------------
    # Sheet 5 – Raw Metrics
    # ------------------------------------------------------------------

    def _sheet_metrics(self, wb: Workbook, data: ReportData) -> None:
        ws = wb.create_sheet("Metrics")

        headers = ["Metric Key", "Value"]
        _header_row(ws, 1, headers)
        ws.freeze_panes = "A2"

        m = data.metrics
        metric_rows = [
            ("quality_gate_status",           m.quality_gate_status),
            ("bugs",                           m.bugs),
            ("reliability_rating",             m.reliability_rating),
            ("reliability_rating_letter",      m.reliability_rating_letter),
            ("vulnerabilities",                m.vulnerabilities),
            ("security_rating",                m.security_rating),
            ("security_rating_letter",         m.security_rating_letter),
            ("code_smells",                    m.code_smells),
            ("sqale_rating",                   m.sqale_rating),
            ("sqale_rating_letter",            m.sqale_rating_letter),
            ("sqale_index_minutes",            m.sqale_index),
            ("sqale_debt_display",             m.technical_debt_display),
            ("sqale_debt_ratio",               m.sqale_debt_ratio),
            ("coverage",                       m.coverage),
            ("duplicated_lines_density",       m.duplicated_lines_density),
            ("duplicated_blocks",              m.duplicated_blocks),
            ("lines",                          m.lines),
            ("ncloc",                          m.ncloc),
            ("files",                          m.files),
            ("functions",                      m.functions),
            ("classes",                        m.classes),
            ("complexity",                     m.complexity),
            ("cognitive_complexity",           m.cognitive_complexity),
            ("comment_lines_density",          m.comment_lines_density),
            ("open_issues",                    m.open_issues),
            ("confirmed_issues",               m.confirmed_issues),
            ("false_positive_issues",          m.false_positive_issues),
            ("wont_fix_issues",                m.wont_fix_issues),
            ("accepted_issues",                m.accepted_issues),
            ("new_bugs",                       m.new_bugs),
            ("new_vulnerabilities",            m.new_vulnerabilities),
            ("new_code_smells",                m.new_code_smells),
            ("new_coverage",                   m.new_coverage),
            ("new_duplicated_lines_density",   m.new_duplicated_lines_density),
        ]

        for row_idx, (key, val) in enumerate(metric_rows, start=2):
            ws.cell(row=row_idx, column=1, value=key).font  = _font(bold=True)
            ws.cell(row=row_idx, column=2, value=val).alignment = _center()
            for col in (1, 2):
                ws.cell(row=row_idx, column=col).border = _border()
            if row_idx % 2 == 0:
                ws.cell(row=row_idx, column=1).fill = _fill(C_ALT_ROW)
                ws.cell(row=row_idx, column=2).fill = _fill(C_ALT_ROW)

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20

    # ------------------------------------------------------------------
    # Helper
    # ------------------------------------------------------------------

    @staticmethod
    def _write_section_title(ws, row: int, col: str, title: str) -> None:
        cell = ws[f"{col}{row}"]
        cell.value = title
        cell.font  = _font(bold=True, color=C_WHITE, size=11)
        cell.fill  = _fill(C_SECONDARY)
        cell.alignment = _left()
